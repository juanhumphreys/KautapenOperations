"""Helpers compartidos para leer Excel con openpyxl."""
from __future__ import annotations
from pathlib import Path
from typing import Any, Iterator
import openpyxl


def open_workbook(path: Path):
    """Abre read-only + data_only (sin fórmulas, sólo valores cacheados)."""
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def iter_rows(ws, start_row: int = 1) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """Itera (n_fila, valores) saltando filas iniciales y devolviendo el índice 1-based."""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < start_row:
            continue
        yield i, row


# Centinelas que aparecen en las celdas por fórmulas rotas.
BROKEN = {"#REF!", "#N/A", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!"}


def is_broken(v: Any) -> bool:
    return isinstance(v, str) and v.strip() in BROKEN


def as_str(v: Any) -> str:
    """str() defensivo: None → ''."""
    if v is None:
        return ""
    return str(v).strip()


def as_float(v: Any) -> float | None:
    """float() defensivo. Devuelve None si la celda no es numérica usable."""
    if v is None or v == "":
        return None
    if is_broken(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def as_code(v: Any) -> str:
    """Códigos de cuenta: a veces vienen como int, a veces str. Normalizar a str sin .0."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()
